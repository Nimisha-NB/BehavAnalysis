import os
import openpyxl

OUTPUT_FILE = "predictions_0.66.xlsx" #whatever excel file name
SHEET_NAME  = 'Light and Shadow' # sheetname
# def process_file_and_store(results, file_name):
def process_file_and_store(results):
    file_exists = os.path.exists(OUTPUT_FILE)

    if file_exists:
        workbook = openpyxl.load_workbook(OUTPUT_FILE)
        if SHEET_NAME in workbook.sheetnames:
            worksheet = workbook[SHEET_NAME]
        else:
            worksheet = workbook.create_sheet(title=SHEET_NAME)
            # worksheet.append(["File Name", "Chunk", "Predicted Label", "Confidence"]) 
            worksheet.append(["Predicted Label", "Confidence"]) 

        # worksheet = workbook.active  # Use the existing sheet
    else:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = SHEET_NAME
        # worksheet.append(["File Name", "Chunk", "Predicted Label", "Confidence"]) 
        worksheet.append(["Predicted Label", "Confidence"])

    for result in results:
        try:
            # text_chunk = result["Text"]
            for prediction in result["Predictions"]:
                worksheet.append([
                    # file_name,
                    # text_chunk[::]
                    prediction["Label"],
                    prediction["Confidence"]
                ])
        except:
            pass

    workbook.save(OUTPUT_FILE)
    print(f"Predictions appended to {OUTPUT_FILE}")
