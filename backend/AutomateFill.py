import openpyxl
import re
from collections import defaultdict

# File paths
input_file = "trainingSet3.xlsx"
output_file = "Updated_Training_Data.xlsx"

# Load the input workbook and sheet
wb_input = openpyxl.load_workbook(input_file)
ws_input = wb_input.active  # Assuming data is in the first sheet

# Dictionary to store grouped data
data_dict = defaultdict(list)
all_data = []  # List to store all rows (formatted and unformatted)

# Function to extract number from the first column
def extract_number(text):
    match = re.match(r"(\d+)", str(text))
    return match.group(1) if match else None

# Function to extract label from the second column
def extract_label(text):
    match = re.match(r"([A-Za-z]+\d*)", str(text))  # Captures A, B, G1, etc.
    return match.group(1) if match else None

try:
# Read input file and process data
    for row in ws_input.iter_rows(min_row=2, values_only=True):  # Skipping header
        number = extract_number(row[0])  # Extract number from column 1
        label = extract_label(row[1])  # Extract letter from column 2
        text = str(row[3]).strip() if row[3] else None  # Text from column 4
except:
    for row in ws_input.iter_rows(min_row=2, values_only=True):
        label = str(row[0])
        number = ""
        text = str(row[1]).strip() if row[1] else None  # Text from column 4


    if number and label and text: 
        full_label = number + label  # Combine number and label (e.g., "5B")
        all_data.append([full_label, text.replace('“', '"').replace('”', '"').replace("‘", "'").replace("’","'")])
    else:  # If data does not match expected format, store as is
        all_data.append(list(row))

# Try loading the output file; create a new one if it doesn't exist
try:
    wb_output = openpyxl.load_workbook(output_file)
    ws_output = wb_output.active
except FileNotFoundError:
    wb_output = openpyxl.Workbook()
    ws_output = wb_output.active
    ws_output.append(["Label", "Text"])  # Adding headers if new file

# Append new data to the existing sheet
for row in all_data:
    ws_output.append(row)

# Sorting entire sheet (excluding headers)
data_to_sort = list(ws_output.iter_rows(min_row=2, values_only=True))
data_to_sort.sort(key=lambda x: (str(x[0]) if x[0] is not None else "", str(x[1]) if x[1] is not None else ""))

# Remove duplicates based on "Text" column
unique_texts = set()
sorted_unique_data = []
for row in data_to_sort:
    if row[1] not in unique_texts:  # Check if text is unique
        sorted_unique_data.append(row)
        unique_texts.add(row[1])

# Write sorted unique data back (without deleting headers)
for i, row in enumerate(sorted_unique_data, start=2):  # Start from row 2 (skip headers)
    for j, value in enumerate(row, start=1):
        ws_output.cell(row=i, column=j, value=value)

# Save the updated file
wb_output.save(output_file)
print(f"Data appended, sorted, and duplicates removed in {output_file}")
